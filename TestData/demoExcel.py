import openpyxl

# this is to load the excel sheet in to variable
book = openpyxl.load_workbook(r"D:\Python\testdata.xlsx")

# this looking for the active sheet in the excel workbook
sheet = book.active

# to select the cell in the excel sheet
cell = sheet.cell(row=1, column=2)
print(cell.value)

# to write the values in the particular cell in the sheet
sheet.cell(row=2, column=2).value = "Uday"
print(sheet.cell(row=2, column=2).value)

# to find the maximum rows and column in the sheet use "max and min"
print(sheet.max_row)
print(sheet.max_column)

# another way to print the values from excelsheet
print(sheet["A5"].value)
print("************************************")
Dict = {}
# using for loop to iterate all the values in the excelsheet
for i in range(1, sheet.max_row+1):
    # print only testcase 2 values
    if sheet["A{}".format(i)].value == "Testcase2":
        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)




