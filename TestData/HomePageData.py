import openpyxl


class HomePageData:

    test_homepage_data = [{"Firstname": "Rahul", "Lastname":"Shetty", "gender":"Female"},
                            {"Firstname": "Uday", "Lastname": "Kumar", "gender": "Male"}]

    '''class irukura method call pananuna namaba object create pananum
    without object if we need to call the method we need to declare that method
    as Static method
    for Non Static method only we have "self" keyword'''
    @staticmethod
    def gettestdata(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(r"D:\Python\testdata.xlsx")
        sheet = book.active
        # using for loop to iterate all the values in the excelsheet
        for i in range(1, sheet.max_row + 1):
            # print only testcase 2 values
            if sheet["A{}".format(i)].value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]